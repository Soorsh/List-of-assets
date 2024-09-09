from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from openpyxl import load_workbook
from datetime import datetime
from bs4 import BeautifulSoup
import pandas as pd
import time
import re
import os

chrome_options = Options()
chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Safari/537.36")
driver_path = 'chromedriver-win64\\chromedriver.exe'
chrome_service = Service(driver_path)
driver = webdriver.Chrome(service=chrome_service, options=chrome_options)
try:
    driver.get('https://trading.finam.ru/profile/MOEX-LKOH')
    print("Страница загружена...")
    try:
        markets_tab_button = WebDriverWait(driver, 35).until(
            EC.presence_of_element_located((By.ID, 'markets-tab-button'))
        )
        markets_tab_button.click()
        print("Кнопка 'Маркет' нажата.")
    except TimeoutException:
        print("Время ожидания истекло. Кнопка 'Маркет' не найдена.")
    try:
        market_group_dropdown = WebDriverWait(driver, 35).until(
            EC.presence_of_element_located((By.ID, 'market-group-dropdown'))
        )
        market_group_dropdown.click()
        print("Кнопка 'Группа рынка' нажата.")
    except TimeoutException:
        print("Время ожидания истекло. Кнопка 'Группа рынка' не найдена.")
    try:
        time.sleep(1)
        stock_and_funds_element = WebDriverWait(driver, 35).until(
            EC.presence_of_element_located((By.XPATH, "//li[contains(text(), 'Акции и фонды')]"))
        )
        stock_and_funds_element.click()
        print("Нажата кнопка 'Акции и фонды'.")
    except TimeoutException:
        print("Время ожидания истекло. Элемент 'Акции и фонды' не найден.")
    try:
        element = WebDriverWait(driver, 35).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'div[data-testid="virtuoso-item-list"]'))
        )
        print("Элемент найден. Ждем 10 секунд до полной загрузки...")
        driver.execute_script("""
            const targetNode = document.querySelector('div[data-testid="virtuoso-item-list"]');
            const config = { childList: true, subtree: true };
            const callback = function(mutationsList) {
                for(const mutation of mutationsList) {
                    if (mutation.type === 'childList') {
                        mutation.addedNodes.forEach(node => {
                            if (node.nodeType === 1) {
                                node.style.height = '1px';
                            }
                        });
                    }
                }
            };
            const observer = new MutationObserver(callback);
            observer.observe(targetNode, config);
            const existingItems = targetNode.children;
            for (let i = 0; i < existingItems.length; i++) {
                existingItems[i].style.height = '1px';
            }
        """)
        time.sleep(10)
    except TimeoutException:
        print("Время ожидания истекло. Элемент 'virtuoso-item-list' не найден.")

    html_content = driver.page_source
    soup = BeautifulSoup(html_content, 'html.parser')
    virtuoso_item_lists = soup.select('div[data-testid="virtuoso-item-list"]')
    formatted_text = ""
    if virtuoso_item_lists:
        print("Элементы найдены.")
        for virtuoso_item_list in virtuoso_item_lists:
            button_elements = virtuoso_item_list.find_all('div', role='button')
            for button in button_elements:
                p_elements = button.find_all('p')
                structured_texts = []
                for p in p_elements:
                    text = p.get_text(strip=True)
                    structured_texts.append(text)
                if structured_texts:
                    formatted_text += ' / '.join(structured_texts) + ' /\n'
        #  print(formatted_text)
    else:
        print("Элементы virtuoso-item-list не найдены.")

    # Удаляем лишние пробелы в начале и конце текста и разбиваем текст на строки
    lines = formatted_text.strip().split("\n")
    data = []

    for line in lines:
        line = line.strip().strip('/').replace('\xa0', '').replace('₽', '')
        line = re.sub(r'\s+', '', line)
        line = line[:line.rfind('/')]
        fields = [field.strip() for field in line.split('/') if field.strip()]
        data.append(fields)
    # print("Поля:", data)

    # Экспорт данных в Excel
    current_date = datetime.now().strftime('%d.%m.%Y')
    df = pd.DataFrame(data, columns=['Компания', 'Код', current_date])
    excel_file = 'Активы.xlsx'

    if os.path.exists(excel_file):
        try:
            # Загружаем существующий файл Excel
            workbook = load_workbook(excel_file)
            sheet = workbook[workbook.sheetnames[0]]

            # Получаем коды из первого столбца Excel
            existing_codes = [cell.value for cell in sheet['B'][1:]]  # Из столбца Код
            existing_companies = [cell.value for cell in sheet['A'][1:]]  # Из столбца Компания

            # Создаем список для упорядоченных данных
            ordered_data = []

            # Упорядочиваем данные по кодам из Excel и добавляем нули для отсутствующих кодов
            for code in existing_codes:
                if code in df['Код'].values:
                    # Если код найден в новом наборе данных, берем цену
                    price = df.loc[df['Код'] == code, current_date].values[0]
                else:
                    # Если код отсутствует в новом наборе данных, присваиваем цену 0
                    price = 0
                ordered_data.append((code, price))

            # Добавляем новые коды и компании, которые есть в df, но отсутствуют в Excel
            new_codes = df[~df['Код'].isin(existing_codes)]
            for _, row in new_codes.iterrows():
                ordered_data.append((row['Код'], row[current_date]))
                # Добавляем компанию, если её нет в существующих компаниях
                if row['Компания'] not in existing_companies:
                    sheet.append([row['Компания'], row['Код'], ''])  # Предварительно добавляем пустую цену

            # Создаем новый DataFrame из упорядоченных данных
            ordered_df = pd.DataFrame(ordered_data, columns=['Код', current_date])

            # Проверяем, существует ли столбец с текущей датой
            column_names = [cell.value for cell in sheet[1]]  # Заголовки из первой строки
            if current_date not in column_names:
                # Если столбца с сегодняшней датой нет, добавляем его
                column_index = len(column_names) + 1
                sheet.cell(row=1, column=column_index, value=current_date)

            else:
                # Если столбец существует, находим его индекс
                column_index = column_names.index(current_date) + 1

            # Вставляем упорядоченные цены в соответствующий столбец
            for row_idx, (code, price) in enumerate(ordered_data, start=2):  # Начинаем со второй строки
                sheet.cell(row=row_idx, column=column_index, value=price)

            # Сохраняем изменения
            workbook.save(excel_file)
            print("Данные успешно обновлены в файле")

        except PermissionError:
            print("Закройте таблицу >:(")
        except Exception as e:
            print(f"Ошибка при чтении файла: {e}")

    else:
        # Если файла нет, создаем его
        df.to_excel(excel_file, index=False)
        print("Данные успешно сохранены в новый файл")
finally:
    driver.quit()